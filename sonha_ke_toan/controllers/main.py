from odoo import http
from odoo.http import request, content_disposition
import io
from datetime import date

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os


current_dir = os.path.dirname(os.path.abspath(__file__))

font_path = os.path.join(current_dir, '..', 'static', 'fonts', 'DejaVuSans.ttf')
font_bold_path = os.path.join(current_dir, '..', 'static', 'fonts', 'DejaVuSans-Bold.ttf')

pdfmetrics.registerFont(TTFont('DejaVu', font_path))
pdfmetrics.registerFont(TTFont('DejaVu-Bold', font_bold_path))


class FieldConfirmController(http.Controller):

    @http.route('/field_confirm/get_fields', type='json', auth='user')
    def get_fields(self, model):
        # trả về danh sách tên field cần confirm (['ma','cap',...])
        return request.env['field.confirm.config'].sudo().get_confirm_fields_for_model(model)

    @http.route("/sonha/get_allowed_models", type="json", auth="user")
    def get_allowed_models(self):
        user = request.env.user
        models = request.env["sonha.phan.quyen"].sudo().search([
            ("NGUOI_DUNG", "=", user.id),
            ("XEM_DM", "=", True),
        ]).mapped("TEN_BANG")
        return models

    @http.route('/field_confirm/get_confirm_fields', type='json', auth='user')
    def get_confirm_fields(self, model):
        fields = request.env['sonha.xac.nhan'].sudo().search([
            ('MODEL_ID.model', '=', model),
            ('REQUIRE_CONFIRM', '=', True),
        ])
        field_names = fields.mapped('FIELD_ID.name')
        return field_names

    def _safe(self, value, default=''):
        return value or default

    def _fmt_vn_date(self, d):
        if not d:
            return ''
        if isinstance(d, str):
            return d
        return f"Ngày {d.day:02d} tháng {d.month:02d} năm {d.year}"


    def _download_phieu_nhap_generic(self, model_name, record_id, default_note='Phiếu nhập mua hàng'):
        record = request.env[model_name].browse(record_id)
        if not record.exists():
            return request.not_found()

        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        logo_x = 20 * mm
        logo_y = height - 35 * mm
        p.setStrokeColor(colors.HexColor('#1f77b4'))
        p.setLineWidth(1)
        p.rect(logo_x, logo_y, 16 * mm, 16 * mm, stroke=1, fill=0)
        p.setFont('DejaVu-Bold', 11)
        p.setFillColor(colors.HexColor('#1f77b4'))
        p.drawCentredString(logo_x + 8 * mm, logo_y + 8 * mm - 4, 'LOGO')
        p.setFillColor(colors.black)

        p.setFont('DejaVu-Bold', 12)
        p.drawString(42 * mm, height - 20 * mm, self._safe(record.DVCS.name, 'CÔNG TY CỔ PHẦN [TÊN CÔNG TY]'))
        p.setFont('DejaVu', 10)
        p.drawString(42 * mm, height - 28 * mm, self._safe(record.DVCS.partner_id.contact_address, 'Địa chỉ: [ĐỊA CHỈ CÔNG TY]'))

        p.setFont('DejaVu-Bold', 11)
        p.drawRightString(width - 20 * mm, height - 20 * mm, 'Mẫu số 01 - VT')
        p.setFont('DejaVu', 10)
        p.drawRightString(width - 20 * mm, height - 28 * mm, '(Ban hành theo QĐ số 15/2006/QĐ-BTC')
        p.setFont('DejaVu', 10)
        p.drawRightString(width - 20 * mm, height - 28 * mm, 'ngày 20/03/2006 của Bộ trưởng BTC)')

        p.setFont('DejaVu-Bold', 18)
        p.drawCentredString(width / 2, height - 45 * mm, 'PHIẾU NHẬP KHO')
        p.setFont('DejaVu', 10)
        p.drawCentredString(width / 2, height - 54 * mm, self._fmt_vn_date(record.NGAY_CT or date.today()))
        p.drawCentredString(width / 2, height - 62 * mm, f"Số: {self._safe(record.CHUNG_TU, f'PN/{record.id}')}")

        y = height - 78 * mm
        p.setFont('DejaVu', 10)
        p.drawString(12 * mm, y, f"Họ và tên người giao hàng: {self._safe(record.KHACH_HANG.TEN, '')}")
        y -= 10 * mm
        p.drawString(12 * mm, y, f"Theo hóa đơn số: {self._safe(record.SO_HD, '')}")
        p.drawRightString(width - 20 * mm, y, self._fmt_vn_date(record.NGAY_HD or record.NGAY_CT))
        y -= 10 * mm
        p.drawString(12 * mm, y, f"Nhập tại kho: {self._safe(record.KHO.TEN, '')}")
        y -= 10 * mm
        p.drawString(12 * mm, y, f"Nội dung: {self._safe(record.GHI_CHU, default_note)}")

        table_top = y - 8 * mm
        lines = record.ACC_SP_D
        data = [
            ['Stt', 'Tên, nhãn hiệu, quy cách,\nphẩm chất vật tư, dụng cụ\nsản phẩm hàng hóa', 'Mã vật tư', 'Đơn vị\ntính', 'Số lượng', '', 'Ghi chú'],
            ['', '', '', '', 'Theo chứng\ntừ', 'Thực nhập', ''],
        ]
        for idx, line in enumerate(lines, start=1):
            hh_name = line.HANG_HOA.TEN_HANG if hasattr(line.HANG_HOA, 'TEN_HANG') else line.HANG_HOA.TEN
            ma_hh = self._safe(getattr(line.HANG_HOA, 'MA_HANG', ''), '')
            dvt = self._safe(getattr(line.HANG_HOA, 'DVT', ''), '')
            data.append([str(idx), self._safe(hh_name, ''), ma_hh, dvt, str(line.SO_LUONG or ''), str(line.SO_LUONG2 or line.SO_LUONG or ''), ''])

        if len(data) == 2:
            data.append(['1', '', '', '', '', '', ''])

        table = Table(data, colWidths=[10 * mm, 67 * mm, 30 * mm, 15 * mm, 22 * mm, 22 * mm, 32 * mm])
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.6, colors.black), ('SPAN', (0, 0), (0, 1)), ('SPAN', (1, 0), (1, 1)),
            ('SPAN', (2, 0), (2, 1)), ('SPAN', (3, 0), (3, 1)), ('SPAN', (4, 0), (5, 0)), ('SPAN', (6, 0), (6, 1)),
            ('FONTNAME', (0, 0), (-1, 1), 'DejaVu-Bold'), ('FONTSIZE', (0, 0), (-1, 1), 8), ('ALIGN', (0, 0), (-1, 1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('FONTNAME', (0, 2), (-1, -1), 'DejaVu'), ('FONTSIZE', (0, 2), (-1, -1), 8),
            ('ALIGN', (0, 2), (0, -1), 'CENTER'), ('ALIGN', (2, 2), (6, -1), 'CENTER'), ('LEFTPADDING', (1, 2), (1, -1), 3),
            ('LINEABOVE', (0, 2), (-1, 2), 0.6, colors.black),]))

        tw, th = table.wrapOn(p, width - 35 * mm, height)
        table.drawOn(p, 6 * mm, table_top - th)

        sign_y = table_top - th - 28 * mm
        p.setFont('DejaVu', 14)
        p.drawRightString(width - 20 * mm, sign_y + 16 * mm, self._fmt_vn_date(record.NGAY_CT or date.today()))

        p.setFont('DejaVu-Bold', 14)
        roles = ['Người lập phiếu', 'Quản đốc', 'Thủ kho', 'Kế toán trưởng']
        x_positions = [35 * mm, 85 * mm, 135 * mm, 185 * mm]
        for x, role in zip(x_positions, roles):
            p.drawCentredString(x, sign_y, role)
            p.setFont('DejaVu', 12)
            p.drawCentredString(x, sign_y - 8 * mm, '(Ký, họ tên)')
            p.setFont('DejaVu-Bold', 14)

        p.showPage(); p.save()
        pdf = buffer.getvalue(); buffer.close()
        return request.make_response(pdf, headers=[('Content-Type', 'application/pdf'),('Content-Disposition', f'attachment; filename="phieu_nhap_{record.id}.pdf"')])

    @http.route('/download/phieu_nhap/<int:record_id>', type='http', auth='user')
    def download_pdf(self, record_id, **kwargs):
        return self._download_phieu_nhap_generic('nl.acc.ap.h', record_id, 'Phiếu nhập mua hàng')

    @http.route('/download/phieu_nhap_nk/<int:record_id>', type='http', auth='user')
    def download_pdf_nk(self, record_id, **kwargs):
        return self._download_phieu_nhap_generic('acc.ap.nk.h', record_id, 'Phiếu nhập mua hàng - Nhập khẩu')

    @http.route('/download/phieu_nhap_tl/<int:record_id>', type='http', auth='user')
    def download_pdf_tl(self, record_id, **kwargs):
        return self._download_phieu_nhap_generic('nl.acc.ap.tl.h', record_id, 'Phiếu nhập hàng bán bị trả lại')

    @http.route('/download/phieu_nhap_sx/<int:record_id>', type='http', auth='user')
    def download_pdf_sx(self, record_id, **kwargs):
        return self._download_phieu_nhap_generic('nl.acc.nk.sx.h', record_id, 'Phiếu nhập kho thành phẩm')

    @http.route('/download/phieu_chuyen_kho/<int:record_id>', type='http', auth='user')
    def download_pdf_chuyen_kho(self, record_id, **kwargs):
        return self._download_phieu_chuyen_kho_generic('nl.acc.ap.ck.h', record_id)

    @http.route('/download/phieu_xuat_vcnb/<int:record_id>', type='http', auth='user')
    def download_pdf_vcnb(self, record_id, **kwargs):
        return self._download_phieu_xuat_vcnb_generic('nl.acc.vcnb.h', record_id)

    @http.route('/download/phieu_xuat_hang_tra_lai/<int:record_id>', type='http', auth='user')
    def download_pdf_xhtl(self, record_id, **kwargs):
        return self._download_phieu_xuat_kho_generic('nl.acc.ap.xhtl.h', record_id, 'Phiếu xuất hàng trả lại')

    @http.route('/download/phieu_xuat_vat_tu/<int:record_id>', type='http', auth='user')
    def download_pdf_pxvt(self, record_id, **kwargs):
        return self._download_phieu_xuat_kho_generic('nl.acc.ap.pxvt.h', record_id, 'Phiếu xuất vật tư')

    @http.route('/download/phieu_ke_toan_hddv/<int:record_id>', type='http', auth='user')
    def download_pdf_hddv(self, record_id, **kwargs):
        return self._download_phieu_ke_toan_generic('nl.acc.hddv.h', record_id)

    @http.route('/download/phieu_ke_toan_hdbh/<int:record_id>', type='http', auth='user')
    def download_pdf_hdbh(self, record_id, **kwargs):
        return self._download_phieu_ke_toan_generic('nl.acc.hdbh.h', record_id)

    @http.route('/download/phieu_chi/<int:record_id>', type='http', auth='user')
    def download_pdf_phieu_chi(self, record_id, **kwargs):
        return self._download_phieu_chi_generic('nl.acc.pc.h', record_id)

    @http.route('/download/phieu_thu/<int:record_id>', type='http', auth='user')
    def download_pdf_phieu_thu(self, record_id, **kwargs):
        return self._download_phieu_thu_generic('nl.acc.pt.h', record_id)

    def _download_phieu_chuyen_kho_generic(self, model_name, record_id):
        record = request.env[model_name].browse(record_id)
        if not record.exists():
            return request.not_found()

        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        p.setFont('DejaVu-Bold', 12)
        p.drawString(35 * mm, height - 20 * mm, self._safe(record.DVCS.name, 'CÔNG TY CỔ PHẦN [TÊN CÔNG TY]'))
        p.setFont('DejaVu', 10)
        p.drawString(35 * mm, height - 28 * mm, self._safe(record.DVCS.partner_id.contact_address, 'Địa chỉ: [ĐỊA CHỈ CÔNG TY]'))
        p.setFont('DejaVu-Bold', 11)
        p.drawRightString(width - 20 * mm, height - 20 * mm, 'Mẫu số 01 - VT')
        p.setFont('DejaVu', 10)
        p.drawRightString(width - 20 * mm, height - 28 * mm, '(Ban hành theo QĐ số 15/2006/QĐ-BTC')
        p.drawRightString(width - 20 * mm, height - 34 * mm, 'ngày 20/03/2006 của Bộ trưởng BTC)')

        p.setFont('DejaVu-Bold', 18)
        p.drawCentredString(width / 2, height - 45 * mm, 'PHIẾU ĐIỀU CHUYỂN')
        p.setFont('DejaVu', 12)
        p.drawCentredString(width / 2, height - 54 * mm, self._fmt_vn_date(record.NGAY_CT or date.today()))
        p.drawCentredString(width / 2, height - 62 * mm, f"Số : {self._safe(record.CHUNG_TU, f'CK/{record.id}')}")

        y = height - 78 * mm
        p.setFont('DejaVu', 10)
        p.drawString(12 * mm, y, f"Họ và tên người giao hàng: {self._safe(record.ONG_BA, '')}")
        y -= 10 * mm
        p.drawString(12 * mm, y, f"Xuất kho : {self._safe(record.KHO.TEN, '')}")
        y -= 10 * mm
        p.drawString(12 * mm, y, f"Nhập tại kho : {self._safe(record.KHOC.TEN, '')}")
        y -= 10 * mm
        p.drawString(12 * mm, y, f"Nội dung : {self._safe(record.GHI_CHU, 'Phiếu chuyển kho')}")

        table_top = y - 8 * mm
        lines = record.ACC_SP_D
        data = [
            ['Stt', 'Tên, nhãn hiệu, quy cách,\nphẩm chất vật tư, dụng cụ\nsản phẩm hàng hóa', 'Mã vật tư', 'Đơn vị\ntính', 'Số lượng', '', 'Ghi chú'],
            ['', '', '', '', 'Theo chứng\ntừ', 'Thực nhập', ''],
        ]
        for idx, line in enumerate(lines, start=1):
            hh_name = line.HANG_HOA.TEN_HANG if hasattr(line.HANG_HOA, 'TEN_HANG') else line.HANG_HOA.TEN
            ma_hh = self._safe(getattr(line.HANG_HOA, 'MA_HANG', ''), '')
            dvt = self._safe(getattr(line.HANG_HOA, 'DVT', ''), '')
            sl = str(line.SO_LUONG or '')
            data.append([str(idx), self._safe(hh_name, ''), ma_hh, dvt, sl, sl, ''])

        if len(data) == 2:
            data.append(['1', '', '', '', '', '', ''])

        table = Table(data, colWidths=[10 * mm, 67 * mm, 30 * mm, 15 * mm, 22 * mm, 22 * mm, 32 * mm])
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.6, colors.black), ('SPAN', (0, 0), (0, 1)), ('SPAN', (1, 0), (1, 1)),
            ('SPAN', (2, 0), (2, 1)), ('SPAN', (3, 0), (3, 1)), ('SPAN', (4, 0), (5, 0)), ('SPAN', (6, 0), (6, 1)),
            ('FONTNAME', (0, 0), (-1, 1), 'DejaVu-Bold'), ('FONTSIZE', (0, 0), (-1, 1), 8), ('ALIGN', (0, 0), (-1, 1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('FONTNAME', (0, 2), (-1, -1), 'DejaVu'), ('FONTSIZE', (0, 2), (-1, -1), 8),
            ('ALIGN', (0, 2), (0, -1), 'CENTER'), ('ALIGN', (2, 2), (6, -1), 'CENTER'), ('LEFTPADDING', (1, 2), (1, -1), 3),
            ('LINEABOVE', (0, 2), (-1, 2), 0.6, colors.black),]))
        tw, th = table.wrapOn(p, width - 35 * mm, height)
        table.drawOn(p, 6 * mm, table_top - th)

        sign_y = table_top - th - 28 * mm
        p.setFont('DejaVu', 14)
        p.drawRightString(width - 20 * mm, sign_y + 16 * mm, self._fmt_vn_date(record.NGAY_CT or date.today()))
        p.setFont('DejaVu-Bold', 14)
        roles = ['Người lập phiếu', 'Quản đốc', 'Thủ kho']
        x_positions = [35 * mm, 105 * mm, 170 * mm]
        for x, role in zip(x_positions, roles):
            p.drawCentredString(x, sign_y, role)
            p.setFont('DejaVu', 12)
            p.drawCentredString(x, sign_y - 8 * mm, '(Ký, họ tên)')
            p.setFont('DejaVu-Bold', 14)

        p.showPage(); p.save()
        pdf = buffer.getvalue(); buffer.close()
        return request.make_response(pdf, headers=[('Content-Type', 'application/pdf'), ('Content-Disposition', f'attachment; filename=\"phieu_chuyen_kho_{record.id}.pdf\"')])

    def _download_phieu_chi_generic(self, model_name, record_id):
        record = request.env[model_name].browse(record_id)
        if not record.exists():
            return request.not_found()

        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        p.setFont('DejaVu-Bold', 12)
        p.drawString(20 * mm, height - 20 * mm, self._safe(record.DVCS.name, 'CÔNG TY CỔ PHẦN [TÊN CÔNG TY]'))
        p.setFont('DejaVu', 10)
        p.drawString(20 * mm, height - 28 * mm, self._safe(record.DVCS.partner_id.contact_address, ''))

        p.setFont('DejaVu-Bold', 11)
        p.drawRightString(width - 20 * mm, height - 20 * mm, 'Mẫu số 02 - TT')
        p.setFont('DejaVu', 10)
        p.drawRightString(width - 20 * mm, height - 28 * mm, '( Ban hành theo QĐ số 15/2006/QĐ-BTC')
        p.drawRightString(width - 20 * mm, height - 34 * mm, 'ngày 20/03/2006 của Bộ trưởng BTC )')

        p.setFont('DejaVu-Bold', 18)
        p.drawCentredString(width / 2, height - 45 * mm, 'PHIẾU CHI')
        p.setFont('DejaVu', 12)
        p.drawCentredString(width / 2, height - 54 * mm, self._fmt_vn_date(record.NGAY_CT or date.today()))

        p.setFont('DejaVu-Bold', 12)
        p.drawRightString(width - 20 * mm, height - 54 * mm, f"Số: {self._safe(record.CHUNG_TU, f'PC/{record.id}')}")
        p.setFont('DejaVu', 11)
        p.drawRightString(width - 20 * mm, height - 62 * mm, f"Nợ: {self._safe(record.MA_TK0 if hasattr(record, 'MA_TK0') else '', '')}")
        p.drawRightString(width - 20 * mm, height - 69 * mm, f"Có: {self._safe(record.MA_TK1 if hasattr(record, 'MA_TK1') else '', '')}")

        y = height - 78 * mm
        p.setFont('DejaVu', 13)
        p.drawString(20 * mm, y, f"Họ tên người nhận tiền: {self._safe(record.ONG_BA, '')}")
        y -= 10 * mm
        p.drawString(20 * mm, y, f"Địa chỉ: {self._safe(record.DC_THUE or record.KHACH_HANG.DIA_CHI if record.KHACH_HANG else '', '')}")
        y -= 10 * mm
        p.drawString(20 * mm, y, f"Lý do chi: {self._safe(record.GHI_CHU, '')}")
        y -= 14 * mm

        total_amount = sum(record.ACC_SP_D.mapped('PS_NO1'))
        p.drawString(20 * mm, y, f"Số tiền: {total_amount:,.0f} đồng")
        y -= 10 * mm
        p.drawString(20 * mm, y, "Viết bằng chữ:")
        y -= 12 * mm
        p.drawString(20 * mm, y, "Kèm theo:")
        p.drawRightString(width - 20 * mm, y, "Chứng từ gốc.")

        sign_y = y - 24 * mm
        p.setFont('DejaVu', 14)
        p.drawRightString(width - 20 * mm, sign_y + 16 * mm, self._fmt_vn_date(record.NGAY_CT or date.today()))
        p.setFont('DejaVu-Bold', 13)
        roles = ['Giám đốc', 'Kế toán trưởng', 'Người lập phiếu', 'Thủ quỹ', 'Người nhận tiền']
        x_positions = [30 * mm, 65 * mm, 100 * mm, 135 * mm, 170 * mm]
        for x, role in zip(x_positions, roles):
            p.drawCentredString(x, sign_y, role)
            p.setFont('DejaVu', 11)
            p.drawCentredString(x, sign_y - 8 * mm, '(Ký, họ tên)')
            p.setFont('DejaVu-Bold', 13)

        p.showPage()
        p.save()
        pdf = buffer.getvalue()
        buffer.close()
        return request.make_response(pdf, headers=[('Content-Type', 'application/pdf'), ('Content-Disposition', f'attachment; filename=\"phieu_chi_{record.id}.pdf\"')])


    def _download_phieu_thu_generic(self, model_name, record_id):
        record = request.env[model_name].browse(record_id)
        if not record.exists():
            return request.not_found()

        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        p.setFont('DejaVu-Bold', 12)
        p.drawString(20 * mm, height - 20 * mm, self._safe(record.DVCS.name, 'CÔNG TY CỔ PHẦN [TÊN CÔNG TY]'))
        p.setFont('DejaVu', 10)
        p.drawString(20 * mm, height - 28 * mm, self._safe(record.DVCS.partner_id.contact_address, ''))

        p.setFont('DejaVu-Bold', 11)
        p.drawRightString(width - 20 * mm, height - 20 * mm, 'Mẫu số 01 - TT')
        p.setFont('DejaVu', 10)
        p.drawRightString(width - 20 * mm, height - 28 * mm, '( Ban hành theo QĐ số 15/2006/QĐ-BTC')
        p.drawRightString(width - 20 * mm, height - 34 * mm, 'ngày 20/03/2006 của Bộ trưởng BTC )')

        p.setFont('DejaVu-Bold', 18)
        p.drawCentredString(width / 2, height - 45 * mm, 'PHIẾU THU')
        p.setFont('DejaVu', 12)
        p.drawCentredString(width / 2, height - 54 * mm, self._fmt_vn_date(record.NGAY_CT or date.today()))

        p.setFont('DejaVu-Bold', 12)
        p.drawRightString(width - 20 * mm, height - 54 * mm, f"Số: {self._safe(record.CHUNG_TU, f'PT/{record.id}')}")
        p.setFont('DejaVu', 11)
        p.drawRightString(width - 20 * mm, height - 62 * mm, f"Nợ: {self._safe(record.MA_TK0 if hasattr(record, 'MA_TK0') else '', '')}")
        p.drawRightString(width - 20 * mm, height - 69 * mm, f"Có: {self._safe(record.MA_TK1 if hasattr(record, 'MA_TK1') else '', '')}")

        y = height - 78 * mm
        p.setFont('DejaVu', 13)
        p.drawString(20 * mm, y, f"Họ tên người nộp tiền: {self._safe(record.ONG_BA, '')}")
        y -= 10 * mm
        p.drawString(20 * mm, y, f"Địa chỉ: {self._safe(record.DC_THUE or record.KHACH_HANG.DIA_CHI if record.KHACH_HANG else '', '')}")
        y -= 10 * mm
        p.drawString(20 * mm, y, f"Lý do thu: {self._safe(record.GHI_CHU, '')}")
        y -= 14 * mm

        total_amount = sum(record.ACC_SP_D.mapped('PS_NO1'))
        p.drawString(20 * mm, y, f"Số tiền: {total_amount:,.0f} đồng")
        y -= 10 * mm
        p.drawString(20 * mm, y, "Viết bằng chữ:")
        y -= 12 * mm
        p.drawString(20 * mm, y, "Kèm theo:")
        p.drawRightString(width - 20 * mm, y, "Chứng từ gốc.")

        sign_y = y - 24 * mm
        p.setFont('DejaVu', 14)
        p.drawRightString(width - 20 * mm, sign_y + 16 * mm, self._fmt_vn_date(record.NGAY_CT or date.today()))
        p.setFont('DejaVu-Bold', 13)
        roles = ['Giám đốc', 'Kế toán trưởng', 'Người lập phiếu', 'Thủ quỹ', 'Người nộp tiền']
        x_positions = [30 * mm, 65 * mm, 100 * mm, 135 * mm, 170 * mm]
        for x, role in zip(x_positions, roles):
            p.drawCentredString(x, sign_y, role)
            p.setFont('DejaVu', 11)
            p.drawCentredString(x, sign_y - 8 * mm, '(Ký, họ tên)')
            p.setFont('DejaVu-Bold', 13)

        p.showPage()
        p.save()
        pdf = buffer.getvalue()
        buffer.close()
        return request.make_response(pdf, headers=[('Content-Type', 'application/pdf'), ('Content-Disposition', f'attachment; filename="phieu_thu_{record.id}.pdf"')])

    def _download_phieu_xuat_vcnb_generic(self, model_name, record_id):
        record = request.env[model_name].browse(record_id)
        if not record.exists():
            return request.not_found()

        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        p.setFont('DejaVu-Bold', 12)
        p.drawString(12 * mm, height - 20 * mm, self._safe(record.DVCS.name, 'CÔNG TY CỔ PHẦN [TÊN CÔNG TY]'))
        p.setFont('DejaVu', 10)
        p.drawString(12 * mm, height - 28 * mm, f"Địa chỉ: {self._safe(record.DVCS.partner_id.contact_address, '')}")
        p.setFont('DejaVu-Bold', 11)
        p.drawRightString(width - 20 * mm, height - 20 * mm, 'Mẫu số 02 - VT')
        p.setFont('DejaVu', 10)
        p.drawRightString(width - 20 * mm, height - 28 * mm, '(Ban hành theo QĐ số 15/2006/QĐ-BTC')
        p.drawRightString(width - 20 * mm, height - 34 * mm, 'ngày 20/03/2006 của Bộ trưởng BTC)')

        p.setFont('DejaVu-Bold', 18)
        p.drawCentredString(width / 2, height - 45 * mm, 'PHIẾU XUẤT KHO KIÊM')
        p.drawCentredString(width / 2, height - 54 * mm, 'BIÊN NHẬN GIAO HÀNG')
        p.setFont('DejaVu', 12)
        p.drawCentredString(width / 2, height - 62 * mm, self._fmt_vn_date(record.NGAY_CT or date.today()))
        p.setFont('DejaVu-Bold', 12)
        p.drawRightString(width - 20 * mm, height - 62 * mm, f"Số: {self._safe(record.CHUNG_TU, f'VCNB/{record.id}')}")

        y = height - 76 * mm
        p.setFont('DejaVu', 10)
        p.drawString(12 * mm, y, f"Tên khách hàng: {self._safe(record.KHACH_HANG.TEN, '')}")
        y -= 8 * mm
        p.drawString(12 * mm, y, f"Địa chỉ khách hàng: {self._safe(record.KHACH_HANG.DIA_CHI, '')}")
        y -= 8 * mm
        p.drawString(12 * mm, y, f"Diễn giải: {self._safe(record.GHI_CHU, '')}")
        y -= 10 * mm
        p.drawString(12 * mm, y, f"Nhân viên giao hàng: {self._safe(record.ONG_BA, '')}")
        p.drawString(85 * mm, y, f"Số xe: {self._safe(record.CTGS, '')}")
        p.drawString(135 * mm, y, f"Số chứng từ giao hàng: {self._safe(record.CHUNG_TU, '')}")
        y -= 8 * mm
        p.drawString(12 * mm, y, f"Xuất tại kho: {self._safe(record.KHO.TEN, '')}")

        table_top = y - 8 * mm
        lines = record.ACC_SP_D
        data = [['Stt', 'Tên, nhãn hiệu, quy cách,\nphẩm chất vật tư (sản phẩm hàng hóa)', 'Mã số', 'Đơn vị\ntính', 'Số lượng', '', 'Đơn giá', 'Thành tiền'],
                ['', '', '', '', 'Yêu cầu', 'Thực xuất', '', '']]
        total = 0
        for idx, line in enumerate(lines, start=1):
            hh_name = line.HANG_HOA.TEN_HANG if hasattr(line.HANG_HOA, 'TEN_HANG') else line.HANG_HOA.TEN
            ma_hh = self._safe(getattr(line.HANG_HOA, 'MA_HANG', ''), '')
            dvt = self._safe(getattr(line.HANG_HOA, 'DVT', ''), '')
            sl = line.SO_LUONG or 0
            dg = line.DON_GIA or 0
            thanh_tien = sl * dg
            total += thanh_tien
            data.append([str(idx), self._safe(hh_name, ''), ma_hh, dvt, str(sl), str(sl), f"{dg:,.0f}" if dg else '', f"{thanh_tien:,.0f}" if thanh_tien else ''])
        data.append(['', 'Cộng:', '', '', '', '', '', f"{total:,.0f}" if total else ''])

        table = Table(data, colWidths=[10 * mm, 58 * mm, 25 * mm, 15 * mm, 18 * mm, 18 * mm, 20 * mm, 24 * mm])
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.6, colors.black), ('SPAN', (0, 0), (0, 1)), ('SPAN', (1, 0), (1, 1)),
            ('SPAN', (2, 0), (2, 1)), ('SPAN', (3, 0), (3, 1)), ('SPAN', (4, 0), (5, 0)),
            ('FONTNAME', (0, 0), (-1, 1), 'DejaVu-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 8), ('ALIGN', (0, 0), (-1, 1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('ALIGN', (0, 2), (0, -1), 'CENTER'), ('LEFTPADDING', (1, 2), (1, -1), 3),
        ]))
        _, th = table.wrapOn(p, width - 25 * mm, height)
        table.drawOn(p, 10 * mm, table_top - th)

        sign_y = table_top - th - 30 * mm
        p.setFont('DejaVu-Bold', 12)
        roles = ['Đơn vị nhận hàng', 'Vận chuyển', 'Thủ kho', 'Người lập phiếu', 'Người giao hàng', 'Thủ trưởng đơn vị']
        x_positions = [22 * mm, 55 * mm, 88 * mm, 120 * mm, 152 * mm, 185 * mm]
        for x, role in zip(x_positions, roles):
            p.drawCentredString(x, sign_y, role)
            p.setFont('DejaVu', 9)
            p.drawCentredString(x, sign_y - 6 * mm, '(ký, ghi rõ họ, tên)')
            p.setFont('DejaVu-Bold', 12)

        p.showPage(); p.save()
        pdf = buffer.getvalue(); buffer.close()
        return request.make_response(pdf, headers=[('Content-Type', 'application/pdf'), ('Content-Disposition', f'attachment; filename=\"phieu_xuat_vcnb_{record.id}.pdf\"')])

    def _download_phieu_xuat_kho_generic(self, model_name, record_id, default_note):
        record = request.env[model_name].browse(record_id)
        if not record.exists():
            return request.not_found()

        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        p.setFont('DejaVu-Bold', 12)
        p.drawString(35 * mm, height - 20 * mm, self._safe(record.DVCS.name, 'CÔNG TY CỔ PHẦN [TÊN CÔNG TY]'))
        p.setFont('DejaVu', 10)
        p.drawString(35 * mm, height - 28 * mm, self._safe(record.DVCS.partner_id.contact_address, 'Địa chỉ: [ĐỊA CHỈ CÔNG TY]'))
        p.setFont('DejaVu-Bold', 11)
        p.drawRightString(width - 20 * mm, height - 20 * mm, 'Mẫu số 02 - VT')
        p.setFont('DejaVu', 10)
        p.drawRightString(width - 20 * mm, height - 28 * mm, '(Ban hành theo QĐ số 15/2006/QĐ-BTC')
        p.drawRightString(width - 20 * mm, height - 34 * mm, 'ngày 20/03/2006 của Bộ trưởng BTC)')

        p.setFont('DejaVu-Bold', 18)
        p.drawCentredString(width / 2, height - 45 * mm, 'PHIẾU XUẤT KHO')
        p.setFont('DejaVu', 12)
        p.drawCentredString(width / 2, height - 54 * mm, self._fmt_vn_date(record.NGAY_CT or date.today()))
        p.drawCentredString(width / 2, height - 62 * mm, f"Số: {self._safe(record.CHUNG_TU, f'PX/{record.id}')}")

        y = height - 78 * mm
        p.setFont('DejaVu', 10)
        p.drawString(12 * mm, y, f"Họ và tên người nhận hàng: {self._safe(record.KHACH_HANG.TEN, '')}")
        y -= 10 * mm
        p.drawString(12 * mm, y, f"Địa chỉ: {self._safe(record.KHACH_HANG.DIA_CHI, '')}")
        y -= 10 * mm
        p.drawString(12 * mm, y, f"Nội dung: {self._safe(record.GHI_CHU, default_note)}")
        y -= 10 * mm
        p.drawString(12 * mm, y, f"Xuất tại kho: {self._safe(record.KHO.TEN, '')}")

        table_top = y - 8 * mm
        lines = record.ACC_SP_D
        data = [['Stt', 'Tên, nhãn hiệu, quy cách,\nphẩm chất vật tư (sản phẩm hàng hóa)', 'Mã số', 'ĐVT', 'SL Xuất', 'SL Nhận', 'Ghi chú']]
        for idx, line in enumerate(lines, start=1):
            hh_name = line.HANG_HOA.TEN_HANG if hasattr(line.HANG_HOA, 'TEN_HANG') else line.HANG_HOA.TEN
            ma_hh = self._safe(getattr(line.HANG_HOA, 'MA_HANG', ''), '')
            dvt = self._safe(getattr(line.HANG_HOA, 'DVT', ''), '')
            data.append([str(idx), self._safe(hh_name, ''), ma_hh, dvt, str(line.SO_LUONG or ''), str(line.SO_LUONG2 or ''), ''])

        if len(data) == 1:
            data.append(['1', '', '', '', '', '', ''])

        table = Table(data, colWidths=[10 * mm, 67 * mm, 30 * mm, 15 * mm, 22 * mm, 22 * mm, 32 * mm])
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.6, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'DejaVu-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (2, 1), (6, -1), 'CENTER'),
            ('LEFTPADDING', (1, 1), (1, -1), 3),
        ]))
        _, th = table.wrapOn(p, width - 35 * mm, height)
        table.drawOn(p, 6 * mm, table_top - th)

        sign_y = table_top - th - 28 * mm
        p.setFont('DejaVu', 14)
        p.drawRightString(width - 20 * mm, sign_y + 16 * mm, self._fmt_vn_date(record.NGAY_CT or date.today()))
        p.setFont('DejaVu-Bold', 14)
        roles = ['Người lập phiếu', 'NV giao hàng/\nTài xế', 'Bảo vệ/\nKiểm soát', 'Người nhận hàng/\nKhách hàng', 'Kế toán/\nThủ kho/\nNgười nhập liệu']
        x_positions = [25 * mm, 60 * mm, 95 * mm, 140 * mm, 185 * mm]
        for x, role in zip(x_positions, roles):
            for i, line in enumerate(role.split('\n')):
                p.drawCentredString(x, sign_y - (i * 7 * mm / 3), line)
            p.setFont('DejaVu', 12)
            p.drawCentredString(x, sign_y - 16 * mm, '(Ký, họ tên)')
            p.setFont('DejaVu-Bold', 14)

        p.showPage(); p.save()
        pdf = buffer.getvalue(); buffer.close()
        return request.make_response(pdf, headers=[('Content-Type', 'application/pdf'), ('Content-Disposition', f'attachment; filename=\"phieu_xuat_kho_{record.id}.pdf\"')])


    def _download_phieu_ke_toan_generic(self, model_name, record_id, title='PHIẾU KẾ TOÁN'):
        record = request.env[model_name].browse(record_id)
        if not record.exists():
            return request.not_found()

        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        p.setFont('DejaVu-Bold', 12)
        p.drawString(35 * mm, height - 20 * mm, self._safe(record.DVCS.name, 'CÔNG TY CỔ PHẦN [TÊN CÔNG TY]'))
        p.setFont('DejaVu', 10)
        p.drawString(35 * mm, height - 28 * mm, self._safe(record.DVCS.partner_id.contact_address, 'Địa chỉ: [ĐỊA CHỈ CÔNG TY]'))

        p.setFont('DejaVu-Bold', 18)
        p.drawCentredString(width / 2, height - 45 * mm, title)
        p.setFont('DejaVu', 12)
        p.drawCentredString(width / 2, height - 54 * mm, self._fmt_vn_date(record.NGAY_CT or date.today()))
        p.drawCentredString(width / 2, height - 62 * mm, f"Số chứng từ : {self._safe(record.CHUNG_TU, f'CT/{record.id}')}")

        table_top = height - 70 * mm
        lines = record.ACC_SP_D
        data = [['STT', 'Diễn giải', 'Ghi nợ', 'Ghi có', 'Loại tiền', 'Tỷ giá', 'Ngoại tệ', 'Số tiền', 'Ghi chú']]

        total = 0
        for idx, line in enumerate(lines, start=1):
            so_tien = line.PS_NO1 or 0
            total += so_tien
            data.append([
                str(idx),
                self._safe(line.GHI_CHU_CT or record.GHI_CHU, ''),
                self._safe(line.MA_TK0, ''),
                self._safe(line.MA_TK1, ''),
                self._safe(record.TIEN_TE.MA, 'VNĐ'),
                f"{record.TY_GIA or 1:.2f}",
                f"{line.TIEN_NTE or 0:,.0f}",
                f"{so_tien:,.0f}",
                ''
            ])

        data.append(['', 'Tổng cộng', '', '', '', '', '', f"{total:,.0f}", ''])
        table = Table(data, colWidths=[10*mm, 70*mm, 14*mm, 14*mm, 16*mm, 16*mm, 20*mm, 22*mm, 22*mm])
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.6, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'DejaVu-Bold'),
            ('FONTNAME', (0, 1), (-1, -2), 'DejaVu'),
            ('FONTNAME', (0, -1), (-1, -1), 'DejaVu-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
            ('ALIGN', (7, 1), (7, -1), 'RIGHT'),
            ('TEXTCOLOR', (1, -1), (7, -1), colors.red),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        tw, th = table.wrapOn(p, width - 20 * mm, height)
        table.drawOn(p, 10 * mm, table_top - th)

        y = table_top - th - 12 * mm
        p.setFont('DejaVu', 11)
        p.drawString(15 * mm, y, f"Đối tượng liên quan: {self._safe(record.KHACH_HANG.TEN, '')}")
        y -= 8 * mm
        p.drawString(15 * mm, y, f"Địa chỉ: {self._safe(record.DC_THUE, '')}")

        sign_y = y - 40 * mm
        p.setFont('DejaVu', 12)
        roles = ['Giám đốc', 'Kế toán trưởng', 'Người lập']
        x_positions = [40 * mm, 105 * mm, 170 * mm]
        for x, role in zip(x_positions, roles):
            p.drawCentredString(x, sign_y, role)
            p.setFont('DejaVu', 10)
            p.drawCentredString(x, sign_y - 7 * mm, '(Ký, họ tên)')
            p.setFont('DejaVu', 12)

        p.showPage(); p.save()
        pdf = buffer.getvalue(); buffer.close()
        return request.make_response(pdf, headers=[('Content-Type', 'application/pdf'), ('Content-Disposition', f'attachment; filename="phieu_ke_toan_{record.id}.pdf"')])

import base64
import re
import zipfile
import importlib.util
import unicodedata
from copy import copy
from html import escape
from urllib.parse import unquote

try:
    import openpyxl
except ImportError:
    openpyxl = None


class DynamicPhieuInController(http.Controller):
    """Download user-uploaded print templates after filling record data."""

    _placeholder_re = re.compile(r"(\{\{\s*([A-Za-z0-9_\.]+)\s*\}\}|\$\{\s*([A-Za-z0-9_\.]+)\s*\})")

    @http.route('/download/dynamic_phieu_in/<int:phieu_id>', type='http', auth='user')
    def download_dynamic_phieu_in(self, phieu_id, model=None, record_id=None, **kwargs):
        phieu = request.env['danh.muc.phieu.in'].sudo().browse(phieu_id)
        if not phieu.exists() or not phieu.temp:
            return request.not_found()

        model_name = unquote(model or '')
        record = request.env[model_name].browse(int(record_id or 0)).exists()
        if not record:
            return request.not_found()
        record.check_access_rights('read')
        record.check_access_rule('read')

        filename = self._download_filename(phieu)
        content = base64.b64decode(phieu.temp)
        rendered = self._render_template(content, filename, record)
        mimetype = self._guess_mimetype(filename, rendered)
        return request.make_response(rendered, headers=[
            ('Content-Type', mimetype),
            ('Content-Length', str(len(rendered))),
            ('Content-Disposition', content_disposition(filename)),
        ])

    def _download_filename(self, phieu):
        filename = phieu.temp_filename or phieu.ten or 'phieu_in'
        if phieu.temp and not filename.lower().endswith('.pdf'):
            content_head = base64.b64decode(phieu.temp)[:5]
            if content_head == b'%PDF-':
                filename = '%s.pdf' % filename.rsplit('.', 1)[0]
        return filename

    def _render_template(self, content, filename, record):
        ext = (filename.rsplit('.', 1)[-1] if '.' in filename else '').lower()
        values = self._record_values(record)
        if ext == 'pdf' or content.startswith(b'%PDF-'):
            return self._render_pdf(content, values)
        if ext == 'docx':
            return self._render_docx(content, values)
        if ext == 'xlsx':
            return self._render_xlsx(content, values)
        # Text-like fallback: useful for csv/html/xml/txt templates.
        try:
            return self._replace_placeholders(content.decode('utf-8'), values).encode('utf-8')
        except UnicodeDecodeError:
            return content

    def _record_values(self, record):
        values = {'id': record.id, 'display_name': record.display_name or ''}
        line_values = []
        for name, field in record._fields.items():
            if field.type == 'one2many':
                line_values.extend(self._line_values(record[name]))
                continue
            if field.type in ('many2many', 'binary'):
                continue
            try:
                value = record[name]
            except Exception:
                continue
            formatted = self._format_value(value)
            values[name] = formatted
            # Người dùng thường đặt tên ô trong template PDF theo nhãn tiếng Việt
            # thay vì technical field name. Lưu thêm nhãn field để có thể tự map
            # khi fill AcroForm (VD: "Ngày hạch toán", "Số chứng từ").
            if getattr(field, 'string', None):
                values[field.string] = formatted

        values.update(self._date_parts(values))
        values['_lines'] = line_values
        if line_values:
            first_line = line_values[0]
            for key, value in first_line.items():
                values.setdefault(key, value)
        return values


    def _line_values(self, lines):
        result = []
        for line in lines:
            item = {}
            for name, field in line._fields.items():
                if field.type in ('one2many', 'many2many', 'binary'):
                    continue
                try:
                    value = line[name]
                except Exception:
                    continue
                formatted = self._format_value(value)
                item[name] = formatted
                if getattr(field, 'string', None):
                    item[field.string] = formatted
            product = item.get('HANG_HOA') or item.get('Hàng hóa') or item.get('SAN_PHAM') or item.get('Thành phẩm') or ''
            item.setdefault('TEN_SAN_PHAM', product)
            item.setdefault('Tên sản phẩm', product)
            item.setdefault('MA_SAN_PHAM', item.get('MA_HANG') or item.get('Mã hàng') or item.get('MA_VT') or item.get('Mã vật tư') or '')
            item.setdefault('SO_LUONG', item.get('SO_LUONG') or item.get('SL') or item.get('Số lượng') or '')
            item.setdefault('DON_GIA', item.get('DON_GIA') or item.get('Đơn giá') or '')
            item.setdefault('THANH_TIEN', item.get('PS_NO1') or item.get('PS_CO1') or item.get('Thành tiền') or '')
            result.append(item)
        return result

    def _date_parts(self, values):
        date_value = values.get('NGAY_CT') or values.get('Ngày CT') or values.get('NGAY_HD') or values.get('Ngày HĐ')
        if not date_value:
            return {}
        parts = date_value.split('/')
        if len(parts) != 3:
            return {}
        return {
            'NGAY': parts[0], 'THANG': parts[1], 'NAM': parts[2],
            'ngay': parts[0], 'thang': parts[1], 'nam': parts[2],
            'Ngày': parts[0], 'Tháng': parts[1], 'Năm': parts[2],
        }

    def _format_value(self, value):
        if not value:
            return ''
        if hasattr(value, '_name'):
            return value.display_name or ''
        if isinstance(value, (date,)):
            return value.strftime('%d/%m/%Y')
        return str(value)

    def _replace_placeholders(self, text, values, xml=False):
        def repl(match):
            key = match.group(2) or match.group(3)
            value = values.get(key, '')
            return escape(value) if xml else value
        text = self._placeholder_re.sub(repl, text)
        date_value = values.get('NGAY_CT') or values.get('Ngày CT') or values.get('NGAY_HD') or values.get('Ngày HĐ')
        if date_value:
            parts = date_value.split('/')
            if len(parts) == 3:
                text = re.sub(r'ngày\s+\d{1,2}\s+tháng\s+\d{1,2}\s+năm\s+\d{4}',
                              'ngày %s tháng %s năm %s' % (parts[0], parts[1], parts[2]), text, flags=re.IGNORECASE)
        return text

    def _render_docx(self, content, values):
        src = io.BytesIO(content)
        dst = io.BytesIO()
        with zipfile.ZipFile(src, 'r') as zin, zipfile.ZipFile(dst, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith('word/') and item.filename.endswith('.xml'):
                    text = data.decode('utf-8')
                    data = self._replace_placeholders(text, values, xml=True).encode('utf-8')
                zout.writestr(item, data)
        return dst.getvalue()

    def _render_pdf(self, content, values):
        # PDF là file nhị phân nên không thể replace text trực tiếp như docx/xlsx.
        # Với template PDF có AcroForm, fill toàn bộ field trong form bằng dữ
        # liệu bản ghi. Field trong PDF có thể được đặt theo technical field name
        # (vd: ma_kh) hoặc nhãn hiển thị tiếng Việt (vd: Mã khách hàng).
        if not importlib.util.find_spec('pypdf'):
            return content
        import pypdf

        reader = pypdf.PdfReader(io.BytesIO(content))
        writer = pypdf.PdfWriter()
        for page in reader.pages:
            writer.add_page(page)

        acroform = reader.trailer.get('/Root') and reader.trailer['/Root'].get('/AcroForm')
        if acroform:
            # Preserve the document-level AcroForm dictionary. Without this, some
            # PDF viewers keep the old appearance stream or drop the form metadata,
            # so the downloaded file looks blank even though /V was written.
            writer._root_object.update({
                pypdf.generic.NameObject('/AcroForm'): writer._add_object(acroform),
            })
            writer.set_need_appearances_writer(True)
            form_values = self._pdf_form_values(reader, values)
            for page in writer.pages:
                try:
                    writer.update_page_form_field_values(page, form_values, auto_regenerate=True)
                except TypeError:
                    writer.update_page_form_field_values(page, form_values)
            self._set_pdf_field_values(writer, form_values, pypdf)

        output = io.BytesIO()
        writer.write(output)
        return output.getvalue()


    def _set_pdf_field_values(self, writer, form_values, pypdf):
        acroform = writer._root_object.get('/AcroForm')
        fields = acroform.get('/Fields') if acroform else None
        if not fields:
            return

        def update_field(field_ref):
            field = field_ref.get_object()
            field_name = field.get('/T')
            if field_name in form_values:
                value = str(form_values.get(field_name) or '')
                field.update({
                    pypdf.generic.NameObject('/V'): pypdf.generic.TextStringObject(value),
                    pypdf.generic.NameObject('/DV'): pypdf.generic.TextStringObject(value),
                })
            for child in field.get('/Kids') or []:
                update_field(child)

        for field_ref in fields:
            update_field(field_ref)

    def _pdf_form_values(self, reader, values):
        fields = reader.get_fields() or {}
        if not fields:
            return values
        normalized_values = {self._normalize_key(key): value for key, value in values.items()}
        form_values = {}
        for field_name in fields:
            value = values.get(field_name)
            if value is None:
                value = normalized_values.get(self._normalize_key(field_name), '')
            form_values[field_name] = value
        return form_values

    def _normalize_key(self, value):
        value = unicodedata.normalize('NFKD', str(value or ''))
        value = ''.join(char for char in value if not unicodedata.combining(char))
        value = re.sub(r'[^a-zA-Z0-9]+', '', value)
        return value.lower()

    def _render_xlsx(self, content, values):
        if not openpyxl:
            return content
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        cell.value = self._replace_placeholders(cell.value, values)
            self._fill_xlsx_lines(sheet, values.get('_lines') or [])
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _fill_xlsx_lines(self, sheet, lines):
        if not lines:
            return
        header_row = None
        column_map = {}
        for row in sheet.iter_rows():
            normalized = [self._normalize_key(cell.value) for cell in row]
            if any('sanpham' in value or 'hanghoa' in value for value in normalized):
                header_row = row[0].row
                for cell in row:
                    key = self._line_key_from_header(cell.value)
                    if key:
                        column_map[cell.column] = key
                break
        if not header_row or not column_map:
            return
        start_row = header_row + 1
        if len(lines) > 1:
            sheet.insert_rows(start_row + 1, len(lines) - 1)
            for col in range(1, sheet.max_column + 1):
                template_cell = sheet.cell(start_row, col)
                for offset in range(1, len(lines)):
                    target = sheet.cell(start_row + offset, col)
                    if template_cell.has_style:
                        target._style = copy(template_cell._style)
                    target.number_format = template_cell.number_format
                    target.alignment = copy(template_cell.alignment)
                    target.font = copy(template_cell.font)
                    target.fill = copy(template_cell.fill)
                    target.border = copy(template_cell.border)
        for index, line in enumerate(lines, start=1):
            row_number = start_row + index - 1
            for col, key in column_map.items():
                value = index if key == '_stt' else line.get(key, '')
                sheet.cell(row_number, col).value = value

    def _line_key_from_header(self, header):
        normalized = self._normalize_key(header)
        if not normalized:
            return None
        if normalized in ('stt', 'sott') or 'stt' in normalized:
            return '_stt'
        if 'sanpham' in normalized or 'hanghoa' in normalized or 'tennhanhieu' in normalized:
            return 'TEN_SAN_PHAM'
        if 'maso' in normalized or 'mavattu' in normalized or 'mahang' in normalized:
            return 'MA_SAN_PHAM'
        if 'donvitinh' in normalized or normalized == 'dvt':
            return 'DVT'
        if 'soluong' in normalized or normalized.startswith('sl'):
            return 'SO_LUONG'
        if 'dongia' in normalized:
            return 'DON_GIA'
        if 'thanhtien' in normalized:
            return 'THANH_TIEN'
        if 'ghichu' in normalized:
            return 'GHI_CHU'
        return None

    def _guess_mimetype(self, filename, content=None):
        if content and content.startswith(b'%PDF-'):
            return 'application/pdf'
        ext = (filename.rsplit('.', 1)[-1] if '.' in filename else '').lower()
        return {
            'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'pdf': 'application/pdf',
        }.get(ext, 'application/octet-stream')
